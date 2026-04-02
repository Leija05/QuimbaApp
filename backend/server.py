from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
from io import BytesIO
import asyncio
import json
import re
import openpyxl

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')
DATA_DIR = ROOT_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
RECORDS_FILE = DATA_DIR / "records.json"
UPLOADS_FILE = DATA_DIR / "uploads.json"
_records_lock = asyncio.Lock()

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def _load_records_sync() -> List[dict]:
    if not RECORDS_FILE.exists():
        return []
    with RECORDS_FILE.open("r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, list) else []


def _save_records_sync(records: List[dict]) -> None:
    with RECORDS_FILE.open("w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)


async def load_records() -> List[dict]:
    return await asyncio.to_thread(_load_records_sync)


async def save_records(records: List[dict]) -> None:
    await asyncio.to_thread(_save_records_sync, records)


def _load_uploads_sync() -> List[dict]:
    if not UPLOADS_FILE.exists():
        return []
    with UPLOADS_FILE.open("r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, list) else []


def _save_uploads_sync(uploads: List[dict]) -> None:
    with UPLOADS_FILE.open("w", encoding="utf-8") as f:
        json.dump(uploads, f, ensure_ascii=False, indent=2)


async def load_uploads() -> List[dict]:
    return await asyncio.to_thread(_load_uploads_sync)


async def save_uploads(uploads: List[dict]) -> None:
    await asyncio.to_thread(_save_uploads_sync, uploads)


def _normalize_excel_cell(cell_value) -> str:
    if cell_value is None:
        return ""
    return str(cell_value).strip().lower()


def _is_summary_row(row_values) -> bool:
    summary_tokens = (
        "total pendiente",
        "total pagado",
        "total general",
        "tot pendiente",
        "tot pagado",
        "tot general",
    )
    row_text = " ".join(_normalize_excel_cell(value) for value in row_values if value is not None)
    return any(token in row_text for token in summary_tokens) or (
        "total" in row_text and ("pendiente" in row_text or "pagado" in row_text or "general" in row_text)
    )


def _find_header_row(sheet) -> int:
    required_headers = {"fecha", "costo", "transportista", "transporte", "servicio", "status", "estado", "total"}
    scan_limit = min(sheet.max_row or 0, 25)
    best_row = 1
    best_score = -1

    for row_index in range(1, scan_limit + 1):
        row_values = [_normalize_excel_cell(cell.value) for cell in sheet[row_index]]
        score = sum(1 for value in row_values if value in required_headers)
        if score > best_score:
            best_score = score
            best_row = row_index

    return best_row


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _parse_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    cleaned = str(value).strip()
    if cleaned in {"", "-"}:
        return 0.0

    cleaned = cleaned.replace("$", "").replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _is_meaningful_record(record: dict) -> bool:
    return any([
        str(record.get("servicio", "")).strip(),
        str(record.get("transportista", "")).strip(),
        float(record.get("costo_t", 0) or 0) > 0,
        float(record.get("costo_l", 0) or 0) > 0,
        float(record.get("saldo_a_favor", 0) or 0) > 0
    ])


# Define Models
class Record(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    fecha: str
    costo_t: float = 0.0
    transportista: str = ""
    servicio: str = ""
    costo_l: float = 0.0
    status: str = "Pendiente"  # "Pendiente" or "Pagado"
    total: float = 0.0
    saldo_a_favor: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class RecordCreate(BaseModel):
    fecha: str
    costo_t: float = 0.0
    transportista: str = ""
    servicio: str = ""
    costo_l: float = 0.0
    status: str = "Pendiente"
    saldo_a_favor: float = 0.0


class RecordUpdate(BaseModel):
    fecha: Optional[str] = None
    costo_t: Optional[float] = None
    transportista: Optional[str] = None
    servicio: Optional[str] = None
    costo_l: Optional[float] = None
    status: Optional[str] = None
    saldo_a_favor: Optional[float] = None


class TotalsResponse(BaseModel):
    total_pendiente: float
    total_pagado: float
    total_costo_l_pendiente: float


class UploadedFileInfo(BaseModel):
    id: str
    filename: str
    uploaded_at: str
    records_count: int


# Routes
@api_router.get("/")
async def root():
    return {"message": "Sistema de Quimbar API"}


@api_router.get("/records", response_model=List[Record])
async def get_records():
    """Get all records"""
    async with _records_lock:
        return await load_records()


@api_router.get("/records/{record_id}", response_model=Record)
async def get_record(record_id: str):
    """Get a single record by ID"""
    async with _records_lock:
        records = await load_records()
    record = next((record for record in records if record.get("id") == record_id), None)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    return record


@api_router.post("/records", response_model=Record)
async def create_record(data: RecordCreate):
    """Create a new record"""
    # Calculate total
    total = data.costo_t + data.costo_l
    
    record = Record(
        fecha=data.fecha,
        costo_t=data.costo_t,
        transportista=data.transportista,
        servicio=data.servicio,
        costo_l=data.costo_l,
        status=data.status,
        total=total,
        saldo_a_favor=data.saldo_a_favor
    )
    
    doc = record.model_dump()
    async with _records_lock:
        records = await load_records()
        records.append(doc)
        await save_records(records)
    return record


@api_router.put("/records/{record_id}", response_model=Record)
async def update_record(record_id: str, data: RecordUpdate):
    """Update an existing record"""
    async with _records_lock:
        records = await load_records()
        record_index = next((i for i, record in enumerate(records) if record.get("id") == record_id), None)

        if record_index is None:
            raise HTTPException(status_code=404, detail="Record not found")

        existing = records[record_index]
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}

        # Recalculate total if costs changed
        costo_t = update_data.get("costo_t", existing.get("costo_t", 0))
        costo_l = update_data.get("costo_l", existing.get("costo_l", 0))
        update_data["total"] = costo_t + costo_l

        updated = {**existing, **update_data}
        records[record_index] = updated
        await save_records(records)
        return updated


@api_router.delete("/records/{record_id}")
async def delete_record(record_id: str):
    """Delete a record"""
    async with _records_lock:
        records = await load_records()
        filtered_records = [record for record in records if record.get("id") != record_id]
        if len(filtered_records) == len(records):
            raise HTTPException(status_code=404, detail="Record not found")
        await save_records(filtered_records)
    return {"message": "Record deleted successfully"}


@api_router.get("/totals", response_model=TotalsResponse)
async def get_totals():
    """Get totals for pending and paid records"""
    async with _records_lock:
        records = await load_records()
    
    total_pendiente = sum(r.get("total", 0) for r in records if r.get("status") == "Pendiente")
    total_pagado = sum(r.get("total", 0) for r in records if r.get("status") == "Pagado")
    total_costo_l_pendiente = sum(r.get("costo_l", 0) for r in records if r.get("status") == "Pendiente")
    
    return TotalsResponse(
        total_pendiente=total_pendiente,
        total_pagado=total_pagado,
        total_costo_l_pendiente=total_costo_l_pendiente
    )


@api_router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Upload Excel file and import records"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)

        # Map expected columns
        column_map = {
            'fecha': ['fecha', 'date'],
            'costo_t': ['costo_t', 'costo t', 'costot', 'costo transporte', 'costo'],
            'transportista': ['transportista', 'transporter', 'carrier', 'transporte'],
            'servicio': ['servicio', 'service', 'descripcion'],
            'costo_l': ['costo_l', 'costo l', 'costol', 'costo local', 'costo'],
            'status': ['status', 'estado', 'estatus', 'statu'],
            'total': ['total', 'tota'],
            'saldo_a_favor': ['saldo_a_favor', 'saldo a favor', 'saldo', 'balance', 'saldoafavo']
        }
        
        def find_column(field_names, headers):
            normalized_headers = [_normalize_header(header) for header in headers]
            normalized_names = [_normalize_header(name) for name in field_names]

            for index, header in enumerate(normalized_headers):
                if not header:
                    continue
                for name in normalized_names:
                    if not name:
                        continue
                    if header == name or header.startswith(name) or name.startswith(header):
                        return index
            return -1

        records_imported = 0
        errors = []
        parsed_records = []
        seen_records = set()

        async with _records_lock:
            for sheet in workbook.worksheets:
                header_row = _find_header_row(sheet)
                headers = [_normalize_excel_cell(cell.value) for cell in sheet[header_row]]
                col_indices = {field: find_column(names, headers) for field, names in column_map.items()}
                normalized_headers = [_normalize_header(header) for header in headers]
                costo_candidates = [i for i, header in enumerate(normalized_headers) if header.startswith("costo")]

                if col_indices["costo_t"] == col_indices["costo_l"] and len(costo_candidates) >= 2:
                    col_indices["costo_t"] = costo_candidates[0]
                    col_indices["costo_l"] = costo_candidates[1]
                elif col_indices["costo_t"] == col_indices["costo_l"]:
                    # Only one COSTO column was detected; force costo_l to 0
                    col_indices["costo_l"] = -1

                if col_indices["costo_t"] < 0 and costo_candidates:
                    col_indices["costo_t"] = costo_candidates[0]

                if col_indices["costo_l"] < 0 and len(costo_candidates) >= 2:
                    col_indices["costo_l"] = costo_candidates[1]

                for row_num, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    try:
                        if all(value is None or str(value).strip() == "" for value in row):
                            continue

                        if _is_summary_row(row):
                            continue

                        fecha = row[col_indices['fecha']] if col_indices['fecha'] >= 0 else ""
                        if isinstance(fecha, datetime):
                            fecha = fecha.strftime('%Y-%m-%d')
                        elif fecha:
                            fecha = str(fecha)
                        else:
                            fecha = datetime.now().strftime('%Y-%m-%d')

                        costo_t = _parse_float(row[col_indices['costo_t']]) if col_indices['costo_t'] >= 0 else 0
                        transportista = str(row[col_indices['transportista']] or "") if col_indices['transportista'] >= 0 else ""
                        servicio = str(row[col_indices['servicio']] or "") if col_indices['servicio'] >= 0 else ""
                        costo_l = _parse_float(row[col_indices['costo_l']]) if col_indices['costo_l'] >= 0 else 0.0
                        status = str(row[col_indices['status']] or "Pendiente") if col_indices['status'] >= 0 else "Pendiente"
                        saldo_a_favor = _parse_float(row[col_indices['saldo_a_favor']]) if col_indices['saldo_a_favor'] >= 0 else 0

                        # Normalize status
                        status = status.strip().capitalize()
                        if status not in ["Pendiente", "Pagado"]:
                            status = "Pendiente"

                        total_excel = _parse_float(row[col_indices['total']]) if col_indices['total'] >= 0 else 0.0
                        # Prefer explicit TOTAL column from the spreadsheet when present.
                        # Fallback keeps legacy behavior for files without TOTAL.
                        total = total_excel if total_excel > 0 else (
                            costo_t if abs(costo_t - costo_l) < 1e-9 else (costo_t + costo_l)
                        )

                        record = Record(
                            fecha=fecha,
                            costo_t=costo_t,
                            transportista=transportista,
                            servicio=servicio,
                            costo_l=costo_l,
                            status=status,
                            total=total,
                            saldo_a_favor=saldo_a_favor
                        )

                        record_doc = record.model_dump()
                        dedupe_key = (
                            str(record_doc.get("fecha", "")).strip().lower(),
                            str(record_doc.get("transportista", "")).strip().lower(),
                            str(record_doc.get("servicio", "")).strip().lower(),
                            round(float(record_doc.get("costo_t", 0) or 0), 2),
                            round(float(record_doc.get("costo_l", 0) or 0), 2),
                            str(record_doc.get("status", "")).strip().lower(),
                            round(float(record_doc.get("saldo_a_favor", 0) or 0), 2),
                        )

                        if _is_meaningful_record(record_doc) and dedupe_key not in seen_records:
                            seen_records.add(dedupe_key)
                            parsed_records.append(record_doc)
                            records_imported += 1
                    except Exception as e:
                        errors.append(f"Sheet {sheet.title} Row {row_num}: {str(e)}")

            await save_records(parsed_records)

            uploads = await load_uploads()
            uploads.append({
                "id": str(uuid.uuid4()),
                "filename": file.filename,
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
                "records": parsed_records
            })
            await save_uploads(uploads)

        return {
            "message": f"Imported {records_imported} records successfully from {len(workbook.worksheets)} sheet(s)",
            "records_imported": records_imported,
            "sheets_detected": len(workbook.worksheets),
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@api_router.delete("/records")
async def delete_all_records():
    """Delete all records"""
    async with _records_lock:
        records = await load_records()
        deleted_count = len(records)
        await save_records([])
    return {"message": f"Deleted {deleted_count} records"}


@api_router.get("/uploads", response_model=List[UploadedFileInfo])
async def get_uploads():
    async with _records_lock:
        uploads = await load_uploads()
    return [
        UploadedFileInfo(
            id=upload.get("id"),
            filename=upload.get("filename", "Sin nombre"),
            uploaded_at=upload.get("uploaded_at", ""),
            records_count=len(upload.get("records", []))
        )
        for upload in sorted(uploads, key=lambda u: u.get("uploaded_at", ""), reverse=True)
    ]


@api_router.post("/uploads/{upload_id}/load")
async def load_upload_into_table(upload_id: str):
    async with _records_lock:
        uploads = await load_uploads()
        upload = next((item for item in uploads if item.get("id") == upload_id), None)
        if not upload:
            raise HTTPException(status_code=404, detail="Upload not found")
        await save_records(upload.get("records", []))
    return {"message": "Archivo cargado en la tabla", "records_loaded": len(upload.get("records", []))}


@api_router.delete("/uploads/{upload_id}")
async def delete_upload(upload_id: str):
    async with _records_lock:
        uploads = await load_uploads()
        filtered_uploads = [upload for upload in uploads if upload.get("id") != upload_id]
        if len(filtered_uploads) == len(uploads):
            raise HTTPException(status_code=404, detail="Upload not found")
        await save_uploads(filtered_uploads)
    return {"message": "Archivo eliminado del historial"}


@api_router.delete("/uploads")
async def delete_all_uploads():
    async with _records_lock:
        uploads = await load_uploads()
        deleted_count = len(uploads)
        await save_uploads([])
    return {"message": f"Deleted {deleted_count} uploads"}


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    return None
